import datetime
import requests
from openpyxl import load_workbook

#改以下四个企业微信参数
corpid = 'ww73dxxxxxxxx'
corpsecret = '4xxxxxxxx'
agentid = '1000002'
touser = 'ningbodong'


def get_access_token(corpid, corpsecret):
    url = f'https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid={corpid}&corpsecret={corpsecret}'
    response = requests.get(url)
    access_token = response.json().get('access_token')
    return access_token


def send_message(access_token, agentid, content):
    url = f'https://qyapi.weixin.qq.com/cgi-bin/message/send?access_token={access_token}'
    data = {
        "touser": touser,
        "msgtype": "text",
        "agentid": agentid,
        "text": {
            "content": content
        },
        "safe": 0
    }
    response = requests.post(url, json=data)
    return response.json()


def check_due_dates(file_path, date_column, days_before_reminder, title_column):
    today = datetime.datetime.now().date()
    wb = load_workbook(file_path)
    ws = wb.active

    due_date_column_index = ord(date_column.upper()) - 65
    title_column_index = ord(title_column.upper()) - 65

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_cell_value = row[due_date_column_index]
        if isinstance(date_cell_value, datetime.datetime):
            due_date = date_cell_value.date()
            reminder_date = due_date - datetime.timedelta(days=days_before_reminder)
            if today >= reminder_date and today <= due_date:
                title = str(row[title_column_index])
                date = due_date.strftime("%Y-%m-%d")
                message_content = f"提醒：任务'{title}'将于{date}到期，请及时处理。"

                access_token = get_access_token(corpid, corpsecret)
                response = send_message(access_token, agentid, message_content)

                if response.get('errcode') == 0:
                    print("提醒推送成功！")
                else:
                    print("提醒推送失败。")


if __name__ == "__main__":
    excel_file = "test.xlsx"  # 文件名或文件路径
    date_column = "H"  # 日期栏（日期格式）
    days_before_reminder = 5  # 提前提醒天数
    title_column = "B"  # 推送标题
    check_due_dates(excel_file, date_column, days_before_reminder, title_column)
