import datetime
import requests
from openpyxl import load_workbook

def send_reminder(title, date):
    base_url = "http://api.chen-dong.com/2BDFGBFNGN46CB7EFD693.send"  #接口链接：参考https://github.com/xpnas/inotify
    url = f"{base_url}/{title}/{date}"
    response = requests.get(url)
    if response.status_code == 200:
        print("提醒推送成功！")
    else:
        print("提醒推送失败。")

def check_due_dates(file_path, date_column, days_before_reminder, title_column):
    today = datetime.datetime.now().date()
    wb = load_workbook(file_path)
    ws = wb.active

    due_date_column_index = ord(date_column.upper()) - 65
    title_column_index = ord(title_column.upper()) - 65

    for row in ws.iter_rows(min_row=2, values_only=True):
        date_cell_value = row[due_date_column_index]
        if isinstance(date_cell_value, datetime.datetime):
            due_date = date_cell_value.date()
            reminder_date = due_date - datetime.timedelta(days=days_before_reminder)
            if today >= reminder_date and today <= due_date:
                title = str(row[title_column_index])
                date = due_date.strftime("%Y-%m-%d")
                send_reminder(title, date)

if __name__ == "__main__":
    excel_file = "test.xlsx" #文件名或文件路径
    date_column = "H"      #日期栏（日期格式）
    days_before_reminder = 5  #提前提醒天数区间
    title_column = "B"   #推送标题
    check_due_dates(excel_file, date_column, days_before_reminder, title_column)
